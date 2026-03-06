from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = 'Students'

# Create headers
ws['A1'] = 'Name'
ws['B1'] = 'Score'

ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)

# Student data - ensure scores are numbers, not strings
students = [
    ('Alice Johnson', 95),
    ('Bob Smith', 87),
    ('Charlie Brown', 92),
    ('Diana Prince', 78),
    ('Ethan Hunt', 85),
    ('Fiona Green', 91),
    ('George Wilson', 76),
    ('Hannah Montana', 88),
    ('Isaac Newton', 99),
    ('Julia Roberts', 83),
    ('Kevin Hart', 79),
    ('Laura Palmer', 94),
    ('Michael Scott', 72),
    ('Nina Simone', 96),
    ('Oliver Queen', 81),
]

# Add data rows - Name in column A, Score as number in column B
for idx, (name, score) in enumerate(students, start=2):
    ws[f'A{idx}'] = name
    ws[f'B{idx}'] = int(score)  # Store as integer

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10

# Save file
wb.save('sample_grades.xlsx')
print('✅ Excel file created successfully: sample_grades.xlsx')
print('Format:')
print('  Column A: Name (text)')
print('  Column B: Score (numbers)')

