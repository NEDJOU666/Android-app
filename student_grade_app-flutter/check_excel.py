from openpyxl import load_workbook

# Load the Excel file to check its structure
wb = load_workbook('sample_grades.xlsx')
ws = wb.active

print("Sheet name:", ws.title)
print("\nFirst 5 rows:")
for idx, row in enumerate(ws.iter_rows(max_row=5, values_only=False), 1):
    print(f"Row {idx}:")
    for col_idx, cell in enumerate(row[:2], 1):
        if cell:
            print(f"  Cell {col_idx}: value={cell.value}, type={type(cell.value)}")
        else:
            print(f"  Cell {col_idx}: None")

print("\n\nWith values_only=True (what the app should use):")
for idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
    print(f"Row {idx}: {row[:2]}")
